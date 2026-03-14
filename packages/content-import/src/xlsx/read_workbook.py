import json
import os
import sys

from openpyxl import load_workbook


def to_json_value(value):
    if value is None:
        return None
    if isinstance(value, (int, float, bool, str)):
        return value
    return str(value)


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: read_workbook.py <path>")

    workbook_path = sys.argv[1]
    workbook = load_workbook(workbook_path, data_only=False)

    document = {
        "sourceFileName": os.path.basename(workbook_path),
        "sheets": []
    }

    for sheet in workbook.worksheets:
        rows = []
        for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            rows.append({
                "rowNumber": row_index,
                "values": [to_json_value(value) for value in row]
            })

        document["sheets"].append({
            "name": sheet.title,
            "maxRow": sheet.max_row,
            "maxColumn": sheet.max_column,
            "mergedRanges": [str(range_ref) for range_ref in sheet.merged_cells.ranges],
            "rows": rows
        })

    print(json.dumps(document, ensure_ascii=False))


if __name__ == "__main__":
    main()
